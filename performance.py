import time
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl import Workbook
import logging
import threading
import device_data
# package = "com.dianshijia.tvlive"
# package1 = "dianshijia"
# devices = "XPL4C19C11001295"


def excel_info(path_file,chart_name1,chart_name2,package,package1,devices):
    wb = Workbook()
    ws=wb.active
    ws.title = "info_usage"
    #ws_men = wb.create_sheet("info_usage")
    ws["A1"] = "Time"
    ws["B1"] = "men Usage"
    ws['C1']="cpu Usage"
    wb.save(path_file)
    while True:
        men=device_data.get_men(package, devices)
        cpu = device_data.get_cpu(devices,package1)
        wb = load_workbook(path_file)
        ws=wb['info_usage']
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=2, value=men)
        ws.cell(row=row_num, column=3, value=cpu)
        nowTime = device_data.now_time()
        logging.info(nowTime)
        ws.cell(row=row_num, column=1, value=nowTime)
        wb.save(path_file)
        if not men:
            break
        #print("获取手机信息{}".format(threading.enumerate()))
        time.sleep(5)

    wb = load_workbook(path_file)
    ws = wb['info_usage']
    #print("开始调用")
    #chart(path_file,chart_name,wb,ws)
    t2=threading.Thread(target=chart, args=(path_file, chart_name1,chart_name2, wb, ws))
    t2.start()
    t2.join()
    #print("获取使用率结束画出图形{}".format(threading.enumerate()))
    #print("任务结束")
    time.sleep(1)


def chart(path_file,chart_name1,chart_name2,wb,ws):
# 加载 Excel 文件


    A_column_data = [cell.value for row in ws.iter_rows(min_row=2, min_col=1, max_col=1)
                                        for cell in row][:-1]


    B_column_data = [cell.value for row in ws.iter_rows(min_row=2, min_col=2, max_col=2)
                                        for cell in row][:-1]

    # 创建 LineChart 图表对象
    chart1 = LineChart()
    chart2= LineChart()


    #设置图表数据
    xdata = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    ydata1 = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    ydata2 = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
    # chart1.add_data(ydata1, titles_from_data=True)
    # chart1.set_categories(xdata)
    chart1.add_data(ydata1, titles_from_data=True)
    chart1.set_categories(xdata)
    chart2.add_data(ydata2, titles_from_data=True)
    chart2.set_categories(xdata)
    # 设置图表属性
    chart1.title = "{}".format(chart_name1)
    chart2.title = "{}".format(chart_name2)
    chart1.x_axis.title = "Time"
    chart1.y_axis.title = "{}".format(chart_name1)
    chart2.x_axis.title = "Time"
    chart2.y_axis.title = "{}".format(chart_name2)

    # 将图表添加到工作表中
    ws.add_chart(chart1, "D2")
    ws.add_chart(chart2, "D20")

    # 保存 Excel 文件
    wb.save(path_file)


# if __name__ == '__main__':
#     path_file='D:\\xiaxy\\performance_data.xlsx'
#     chart_name='men'
#     #excel_men(path_file,chart_name)

